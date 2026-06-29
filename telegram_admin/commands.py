import asyncio
import logging
import re
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from telegram import Update
from telegram.ext import CallbackContext, ConversationHandler

from database.repository import DatabaseRepository

from . import keyboards

logger = logging.getLogger(__name__)

CHOOSING_PERIOD, AWAITING_CUSTOM_RANGE = range(2)

COMMANDS_TEXT = (
    "Доступные команды:\n"
    "/start — это сообщение\n"
    "/chats — список отслеживаемых чатов\n"
    "/export — выбрать чат и период, получить выгрузку\n"
    "/cancel — отменить текущую операцию"
)


def restricted(func):
    """
    admin_ids теперь берётся из bot_data (его кладёт туда
    build_application в bot.py), а не из прямого
    `from core.config import ADMIN_IDS` — commands.py больше ничего не
    знает про core.config и его можно тестировать отдельно.
    """

    @wraps(func)
    async def wrapper(update: Update, context: CallbackContext, *args, **kwargs):
        admin_ids = context.bot_data.get("admin_ids", set())
        user_id = update.effective_user.id if update.effective_user else None

        if user_id not in admin_ids:
            logger.warning("ADMIN BOT: отказано в доступе, user_id=%s", user_id)

            if update.message:
                await update.message.reply_text("⛔ У вас нет доступа к этому боту.")
            elif update.callback_query:
                await update.callback_query.answer("Доступ запрещён", show_alert=True)

            return ConversationHandler.END

        return await func(update, context, *args, **kwargs)

    return wrapper


async def _run_db(context: CallbackContext, method_name: str, *args, **kwargs):
    """
    Каждый вызов открывает свою собственную, короткоживущую Session (через
    DatabaseRepository(session_factory)) и закрывает её сразу после
    выполнения метода — внутри одного потока из пула asyncio.to_thread.

    Раньше bot_data хранил один общий DatabaseRepository (одну Session) на
    всё время жизни бота, и она использовалась из нескольких разных
    потоков по очереди (поток admin-bot, плюс рабочие потоки
    asyncio.to_thread). Сами методы репозитория при этом возвращали ORM
    объекты (Message и т.п.), у которых код в commands.py потом читал
    отношения (msg.sender, msg.attachments...) уже после возврата из
    to_thread — то есть прямо в event loop'е бота, синхронно, что и
    блокировало его, и было нестабильно. Теперь методы репозитория для
    экспорта (см. build_export_rows) сами разворачивают все связанные
    данные в обычные dict, пока Session ещё открыта, — а сюда наружу
    возвращаются уже простые структуры, без обращения к БД.
    """
    session_factory = context.bot_data["session_factory"]

    def _call():
        db = DatabaseRepository(session_factory)
        try:
            return getattr(db, method_name)(*args, **kwargs)
        finally:
            db.close()

    return await asyncio.to_thread(_call)


async def start(update: Update, context: CallbackContext) -> None:
    """
    /start доступен всем без ограничения: пользователь сразу видит, есть
    у него доступ к боту или нет, и что делать дальше — а не только
    "нет доступа" в ответ на команду, которой он не может воспользоваться.
    """
    admin_ids = context.bot_data.get("admin_ids", set())
    user_id = update.effective_user.id if update.effective_user else None

    if user_id in admin_ids:
        access_line = "✅ У вас есть доступ к этому боту."
    else:
        access_line = (
            "⛔ У вас нет доступа к этому боту.\n"
            "Если он вам нужен — обратитесь к администратору."
        )

    await update.message.reply_text(
        "Привет! Я бот для выгрузки переписки.\n\n"
        f"{access_line}\n\n"
        f"{COMMANDS_TEXT}"
    )


@restricted
async def list_chats(update: Update, context: CallbackContext) -> None:
    chats = await _run_db(context, "list_chats")

    if not chats:
        await update.message.reply_text("Чатов в базе пока нет.")
        return

    text = "Чаты в базе:\n" + "\n".join(
        f"• {chat.title or chat.telegram_chat_id}" for chat in chats
    )
    await update.message.reply_text(text)


@restricted
async def export_start(update: Update, context: CallbackContext) -> int:
    logger.info("ADMIN BOT: export_start вызван, user_id=%s", update.effective_user.id if update.effective_user else None)
    chats = await _run_db(context, "list_chats")

    if not chats:
        await update.message.reply_text("Чатов в базе пока нет.")
        return ConversationHandler.END

    await update.message.reply_text(
        "Выберите чат:", reply_markup=keyboards.chats_keyboard(chats)
    )
    return CHOOSING_PERIOD


@restricted
async def choose_chat(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()
    logger.info(
        "ADMIN BOT: choose_chat вызван, data=%s, user_id=%s",
        query.data, update.effective_user.id if update.effective_user else None,
    )

    try:
        telegram_chat_id = int(query.data.split(":", 1)[1])
    except (IndexError, ValueError):
        logger.exception("ADMIN BOT: некорректный callback_data в choose_chat: %r", query.data)
        await query.edit_message_text("Ошибка выбора чата. Начните заново: /export")
        return ConversationHandler.END

    context.user_data["export_chat_id"] = telegram_chat_id

    await query.edit_message_text(
        "Выберите период:", reply_markup=keyboards.periods_keyboard()
    )
    return CHOOSING_PERIOD


@restricted
async def choose_period(update: Update, context: CallbackContext) -> int:
    query = update.callback_query
    await query.answer()
    logger.info(
        "ADMIN BOT: choose_period вызван, data=%s, user_id=%s",
        query.data, update.effective_user.id if update.effective_user else None,
    )

    try:
        code = query.data.split(":", 1)[1]
    except IndexError:
        logger.exception("ADMIN BOT: некорректный callback_data в choose_period: %r", query.data)
        await query.edit_message_text("Ошибка выбора периода. Начните заново: /export")
        return ConversationHandler.END

    if code == keyboards.PERIOD_CUSTOM:
        await query.edit_message_text(
            "Введите период в формате:\n2026-06-01 2026-06-25"
        )
        return AWAITING_CUSTOM_RANGE

    now = datetime.utcnow()

    if code == keyboards.PERIOD_TODAY:
        date_from = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif code == keyboards.PERIOD_WEEK:
        date_from = now - timedelta(days=7)
    else:  # PERIOD_MONTH
        date_from = now - timedelta(days=30)

    await query.edit_message_text("Готовлю выгрузку…")
    await _send_export(update, context, date_from, now, _filename_range_label(date_from, now))
    return ConversationHandler.END


@restricted
async def custom_range_received(update: Update, context: CallbackContext) -> int:
    # Доп. защита помимо фильтра в bot.py (filters.UpdateType.MESSAGE):
    # если сюда всё же придёт update без обычного нового message (правка,
    # служебное сообщение и т.п.), не падаем с AttributeError, а просто
    # остаёмся в этом же состоянии и просим прислать обычное сообщение.
    if update.message is None or not update.message.text:
        return AWAITING_CUSTOM_RANGE

    try:
        raw_from, raw_to = update.message.text.split()
        date_from = datetime.strptime(raw_from, "%Y-%m-%d")
        date_to = datetime.strptime(raw_to, "%Y-%m-%d") + timedelta(days=1)
    except ValueError:
        await update.message.reply_text(
            "Не понял формат. Пришлите так: 2026-06-01 2026-06-25"
        )
        return AWAITING_CUSTOM_RANGE

    await update.message.reply_text("Готовлю выгрузку…")
    await _send_export(
        update, context, date_from, date_to,
        _filename_range_label(date_from, date_to - timedelta(seconds=1)),
    )
    return ConversationHandler.END


def _safe_filename_part(text: str) -> str:
    text = re.sub(r"[^\w\-]+", "_", text or "", flags=re.UNICODE)
    return text.strip("_") or "export"


def _filename_range_label(date_from, date_to) -> str:
    """
    Раньше в имя файла попадала человекочитаемая метка периода
    ("Последние 7 дней", "Последние 30 дней") — по ней нельзя понять,
    какие именно даты и время попали в выгрузку. Теперь — реальные
    границы периода, с точностью до секунды, как они и используются
    в запросе к БД (Message.tg_date >= date_from AND <= date_to).
    """
    fmt = "%Y-%m-%d_%H-%M-%S"
    return f"{date_from.strftime(fmt)}_{date_to.strftime(fmt)}"


def _format_dt(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(
            max(length + 2, 10), 60
        )


def _build_export_workbook(data: dict) -> BytesIO:
    """
    Три листа: сообщения, вложения, реакции — то есть всё, что лежит в
    БД за выбранный период, а не только текст сообщений, как раньше в CSV.
    """
    wb = Workbook()

    ws_msgs = wb.active
    ws_msgs.title = "Сообщения"
    ws_msgs.append(
        ["ID сообщения", "Дата", "Отправитель", "Текст", "Удалено", "Версия", "Ответ на ID"]
    )
    for row in data["messages"]:
        ws_msgs.append([
            row["telegram_message_id"],
            _format_dt(row["tg_date"]),
            row["sender"],
            row["text"],
            "Да" if row["is_deleted"] else "Нет",
            row["version"],
            row["reply_to"] or "",
        ])

    ws_files = wb.create_sheet("Вложения")
    ws_files.append(
        ["ID сообщения", "Дата сообщения", "Тип", "Имя файла", "MIME", "Размер (байт)", "Путь"]
    )
    for row in data["attachments"]:
        ws_files.append([
            row["telegram_message_id"],
            _format_dt(row["tg_date"]),
            row["file_type"],
            row["file_name"] or "",
            row["mime_type"] or "",
            row["size"] or "",
            row["local_path"] or "",
        ])

    ws_reactions = wb.create_sheet("Реакции")
    ws_reactions.append(
        ["ID сообщения", "Дата сообщения", "Эмодзи", "Custom emoji ID", "Отправитель", "Дата реакции", "Активна"]
    )
    for row in data["reactions"]:
        ws_reactions.append([
            row["telegram_message_id"],
            _format_dt(row["tg_date"]),
            row["emoji"] or "",
            row["custom_emoji_id"] or "",
            row["sender"],
            _format_dt(row["reaction_date"]),
            "Добавлена" if row["event"] == "added" else "Удалена",
        ])

    for ws in (ws_msgs, ws_files, ws_reactions):
        _autosize_columns(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def _send_export(
    update: Update, context: CallbackContext, date_from, date_to, period_label: str
) -> None:
    telegram_chat_id = context.user_data.get("export_chat_id")

    # include_deleted=True: выгрузка должна показывать "всё, что в БД
    # за этот период", включая удалённые сообщения (они помечены
    # отдельной колонкой "Удалено", а не пропадают совсем).
    data = await _run_db(
        context, "build_export_rows", telegram_chat_id, date_from, date_to, True
    )

    chat_id = update.effective_chat.id
    bot = context.bot

    if data is None:
        await bot.send_message(chat_id, "Такой чат не найден в базе.")
        return

    if not (data["messages"] or data["attachments"] or data["reactions"]):
        await bot.send_message(chat_id, "За этот период данных не найдено.")
        return

    workbook = _build_export_workbook(data)

    filename = (
        f"export_{_safe_filename_part(data['chat_title'])}_"
        f"{_safe_filename_part(period_label)}.xlsx"
    )

    await bot.send_document(
        chat_id,
        document=workbook,
        filename=filename,
        caption=(
            f"Сообщений: {len(data['messages'])}, "
            f"вложений: {len(data['attachments'])}, "
            f"реакций: {len(data['reactions'])}"
        ),
    )


async def cancel(update: Update, context: CallbackContext) -> int:
    await update.message.reply_text("Отменено.")
    return ConversationHandler.END